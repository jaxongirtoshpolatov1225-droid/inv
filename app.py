from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import qrcode
import io
import base64
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your-secret-key-here'

db = SQLAlchemy(app)
CORS(app)

# Ma'lumotlar bazasi modellari
class Organization(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    has_floors = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    floors = db.relationship('Floor', backref='organization', lazy=True, cascade='all, delete-orphan')
    rooms = db.relationship('Room', backref='organization', lazy=True, cascade='all, delete-orphan')

class Floor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    organization_id = db.Column(db.Integer, db.ForeignKey('organization.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    rooms = db.relationship('Room', backref='floor', lazy=True, cascade='all, delete-orphan')

class Room(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    organization_id = db.Column(db.Integer, db.ForeignKey('organization.id'), nullable=False)
    floor_id = db.Column(db.Integer, db.ForeignKey('floor.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    equipment = db.relationship('Equipment', backref='room', lazy=True, cascade='all, delete-orphan')

class Equipment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    inv_code = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(200), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    brand = db.Column(db.String(100))
    model = db.Column(db.String(100))
    serial_number = db.Column(db.String(100))
    color = db.Column(db.String(50))
    status = db.Column(db.String(50), default='Active')
    description = db.Column(db.Text)
    room_id = db.Column(db.Integer, db.ForeignKey('room.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Transfer tarixi bilan bog'lanish
    transfer_history = db.relationship('TransferHistory', backref='equipment', lazy=True, cascade='all, delete-orphan')

class TransferHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'), nullable=False)
    from_room_id = db.Column(db.Integer, db.ForeignKey('room.id'), nullable=False)
    to_room_id = db.Column(db.Integer, db.ForeignKey('room.id'), nullable=False)
    old_inv_code = db.Column(db.String(50), nullable=False)
    new_inv_code = db.Column(db.String(50), nullable=False)
    transfer_date = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text)
    
    # Xonalar bilan bog'lanish
    from_room = db.relationship('Room', foreign_keys=[from_room_id], backref='transfers_from')
    to_room = db.relationship('Room', foreign_keys=[to_room_id], backref='transfers_to')

# Asosiy sahifa
@app.route('/')
def index():
    organizations = Organization.query.all()
    return render_template('index.html', organizations=organizations)

# Tashkilot qo'shish
@app.route('/add_organization', methods=['POST'])
def add_organization():
    data = request.get_json()
    name = data.get('name')
    has_floors = data.get('has_floors', False)
    
    if not name:
        return jsonify({'error': 'Tashkilot nomi kiritilishi shart'}), 400
    
    organization = Organization(name=name, has_floors=has_floors)
    db.session.add(organization)
    db.session.commit()
    
    return jsonify({'success': True, 'organization_id': organization.id})

# Tashkilot sahifasi
@app.route('/organization/<int:org_id>')
def organization_page(org_id):
    organization = Organization.query.get_or_404(org_id)
    if organization.has_floors:
        floors = Floor.query.filter_by(organization_id=org_id).all()
        return render_template('organization_with_floors.html', organization=organization, floors=floors)
    else:
        rooms = Room.query.filter_by(organization_id=org_id, floor_id=None).all()
        return render_template('organization_direct_rooms.html', organization=organization, rooms=rooms)

# Qavat qo'shish
@app.route('/add_floor', methods=['POST'])
def add_floor():
    data = request.get_json()
    name = data.get('name')
    organization_id = data.get('organization_id')
    
    if not name or not organization_id:
        return jsonify({'error': 'Qavat nomi va tashkilot ID kiritilishi shart'}), 400
    
    floor = Floor(name=name, organization_id=organization_id)
    db.session.add(floor)
    db.session.commit()
    
    return jsonify({'success': True, 'floor_id': floor.id})

# Qavat sahifasi
@app.route('/floor/<int:floor_id>')
def floor_page(floor_id):
    floor = Floor.query.get_or_404(floor_id)
    rooms = Room.query.filter_by(floor_id=floor_id).all()
    return render_template('floor.html', floor=floor, rooms=rooms)

# Xona qo'shish
@app.route('/add_room', methods=['POST'])
def add_room():
    data = request.get_json()
    name = data.get('name')
    organization_id = data.get('organization_id')
    floor_id = data.get('floor_id')
    
    if not name or not organization_id:
        return jsonify({'error': 'Xona nomi va tashkilot ID kiritilishi shart'}), 400
    
    room = Room(name=name, organization_id=organization_id, floor_id=floor_id)
    db.session.add(room)
    db.session.commit()
    
    return jsonify({'success': True, 'room_id': room.id})

# Xona sahifasi
@app.route('/room/<int:room_id>')
def room_page(room_id):
    room = Room.query.get_or_404(room_id)
    equipment = Equipment.query.filter_by(room_id=room_id).all()
    return render_template('room.html', room=room, equipment=equipment)

# Jihoz qo'shish
@app.route('/add_equipment', methods=['POST'])
def add_equipment():
    data = request.get_json()
    
    # Inventarizatsiya kodi yaratish
    org = Organization.query.get(data['organization_id'])
    room = Room.query.get(data['room_id'])
    equipment_count = Equipment.query.filter_by(room_id=data['room_id']).count() + 1
    
    inv_code = f"{org.name[:3].upper()}-{room.name[:3].upper()}-{equipment_count:04d}"
    
    equipment = Equipment(
        inv_code=inv_code,
        name=data['name'],
        category=data['category'],
        brand=data.get('brand'),
        model=data.get('model'),
        serial_number=data.get('serial_number'),
        color=data.get('color'),
        status=data.get('status', 'Active'),
        description=data.get('description'),
        room_id=data['room_id']
    )
    
    db.session.add(equipment)
    db.session.commit()
    
    return jsonify({'success': True, 'equipment_id': equipment.id, 'inv_code': inv_code})

# QR kod yaratish
@app.route('/generate_qr/<int:equipment_id>')
def generate_qr(equipment_id):
    equipment = Equipment.query.get_or_404(equipment_id)
    
    # QR kod matni (to'liq ma'lumotlar)
    qr_text = f"""TASHKILOT: {equipment.room.organization.name}
QAVAT: {equipment.room.floor.name if equipment.room.floor else 'Yo\'q'}
XONA: {equipment.room.name}
JIHOZ: {equipment.name}
KATEGORIYA: {equipment.category}
INV KODE: {equipment.inv_code}
BREND: {equipment.brand or 'N/A'}
MODEL: {equipment.model or 'N/A'}
SERIYA: {equipment.serial_number or 'N/A'}
RANG: {equipment.color or 'N/A'}
HOLAT: {equipment.status}
TAVSIF: {equipment.description or 'N/A'}"""
    
    # QR kod yaratish (40x30mm qog'oz uchun optimallashtirilgan)
    qr = qrcode.QRCode(
        version=3,           # Kattaroq versiya (ko'p ma'lumot uchun)
        box_size=2,          # Kichikroq box_size (ko'p ma'lumot uchun)
        border=1             # Minimal border
    )
    qr.add_data(qr_text)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Rasmni base64 formatiga o'tkazish
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    img_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    return jsonify({'qr_code': img_base64, 'inv_code': equipment.inv_code})

# Xonadagi barcha jihozlar uchun QR kodlar yaratish
@app.route('/generate_room_qr_codes/<int:room_id>')
def generate_room_qr_codes(room_id):
    room = Room.query.get_or_404(room_id)
    equipment_list = Equipment.query.filter_by(room_id=room_id).all()
    
    if not equipment_list:
        return jsonify({'success': False, 'message': 'Xonada jihoz yo\'q'})
    
    qr_codes = []
    
    for equipment in equipment_list:
        # QR kod matni (to'liq ma'lumotlar)
        qr_text = f"""TASHKILOT: {equipment.room.organization.name}
QAVAT: {equipment.room.floor.name if equipment.room.floor else 'Yo\'q'}
XONA: {equipment.room.name}
JIHOZ: {equipment.name}
KATEGORIYA: {equipment.category}
INV KODE: {equipment.inv_code}
BREND: {equipment.brand or 'N/A'}
MODEL: {equipment.model or 'N/A'}
SERIYA: {equipment.serial_number or 'N/A'}
RANG: {equipment.color or 'N/A'}
HOLAT: {equipment.status}
TAVSIF: {equipment.description or 'N/A'}"""
        
        # QR kod yaratish (40x30mm qog'oz uchun optimallashtirilgan)
        qr = qrcode.QRCode(
            version=3,           # Kattaroq versiya (ko'p ma'lumot uchun)
            box_size=2,          # Kichikroq box_size (ko'p ma'lumot uchun)
            border=1             # Minimal border
        )
        qr.add_data(qr_text)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Rasmni base64 formatiga o'tkazish
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        qr_codes.append({
            'equipment_id': equipment.id,
            'inv_code': equipment.inv_code,
            'name': equipment.name,
            'qr_code': img_base64
        })
    
    return jsonify({
        'success': True,
        'room_name': room.name,
        'equipment_count': len(qr_codes),
        'qr_codes': qr_codes
    })

# Jihoz ma'lumotlarini olish
@app.route('/equipment/<int:equipment_id>')
def get_equipment(equipment_id):
    equipment = Equipment.query.get_or_404(equipment_id)
    return jsonify({
        'id': equipment.id,
        'inv_code': equipment.inv_code,
        'name': equipment.name,
        'category': equipment.category,
        'brand': equipment.brand,
        'model': equipment.model,
        'serial_number': equipment.serial_number,
        'color': equipment.color,
        'status': equipment.status,
        'description': equipment.description,
        'room_name': equipment.room.name,
        'floor_name': equipment.room.floor.name if equipment.room.floor else None,
        'organization_name': equipment.room.organization.name
    })

# Jihoz ma'lumotlarini yangilash
@app.route('/update_equipment/<int:equipment_id>', methods=['PUT'])
def update_equipment(equipment_id):
    equipment = Equipment.query.get_or_404(equipment_id)
    data = request.get_json()
    
    # Ma'lumotlarni yangilash
    equipment.name = data.get('name', equipment.name)
    equipment.category = data.get('category', equipment.category)
    equipment.brand = data.get('brand', equipment.brand)
    equipment.model = data.get('model', equipment.model)
    equipment.serial_number = data.get('serial_number', equipment.serial_number)
    equipment.color = data.get('color', equipment.color)
    equipment.status = data.get('status', equipment.status)
    equipment.description = data.get('description', equipment.description)
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Jihoz ma\'lumotlari yangilandi'})

# Jihozni o'chirish
@app.route('/delete_equipment/<int:equipment_id>', methods=['DELETE'])
def delete_equipment(equipment_id):
    equipment = Equipment.query.get_or_404(equipment_id)
    db.session.delete(equipment)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Jihoz o\'chirildi'})

# Tashkilotni o'chirish
@app.route('/delete_organization/<int:org_id>', methods=['DELETE'])
def delete_organization(org_id):
    organization = Organization.query.get_or_404(org_id)
    
    # Tashkilotdagi barcha xonalarni olish
    rooms = Room.query.filter_by(organization_id=org_id).all()
    
    # Har bir xona uchun transfer_history ni o'chirish
    for room in rooms:
        TransferHistory.query.filter(
            (TransferHistory.from_room_id == room.id) | 
            (TransferHistory.to_room_id == room.id)
        ).delete()
    
    # Keyin tashkilotni o'chirish (qavatlar va xonlar cascade orqali o'chiriladi)
    db.session.delete(organization)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Tashkilot o\'chirildi'})

# Qavatni o'chirish
@app.route('/delete_floor/<int:floor_id>', methods=['DELETE'])
def delete_floor(floor_id):
    floor = Floor.query.get_or_404(floor_id)
    
    # Qavatdagi barcha xonalarni olish
    rooms = Room.query.filter_by(floor_id=floor_id).all()
    
    # Har bir xona uchun transfer_history ni o'chirish
    for room in rooms:
        TransferHistory.query.filter(
            (TransferHistory.from_room_id == room.id) | 
            (TransferHistory.to_room_id == room.id)
        ).delete()
    
    # Keyin qavatni o'chirish (xonlar cascade orqali o'chiriladi)
    db.session.delete(floor)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Qavat o\'chirildi'})

# Xonadagi jihozlar sonini tekshirish
@app.route('/check_room_equipment/<int:room_id>')
def check_room_equipment(room_id):
    room = Room.query.get_or_404(room_id)
    equipment_count = Equipment.query.filter_by(room_id=room_id).count()
    
    return jsonify({
        'has_equipment': equipment_count > 0,
        'equipment_count': equipment_count,
        'room_name': room.name
    })

# Xonani o'chirish (jihozlar bilan birga)
@app.route('/delete_room_with_equipment/<int:room_id>', methods=['DELETE'])
def delete_room_with_equipment(room_id):
    room = Room.query.get_or_404(room_id)
    
    # Avval transfer_history jadvalidagi ma'lumotlarni o'chirish
    TransferHistory.query.filter(
        (TransferHistory.from_room_id == room_id) | 
        (TransferHistory.to_room_id == room_id)
    ).delete()
    
    # Keyin xonani o'chirish (jihozlar cascade orqali o'chiriladi)
    db.session.delete(room)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Xona va barcha jihozlar o\'chirildi'})

# Xonani o'chirish (jihozlarsiz)
@app.route('/delete_room/<int:room_id>', methods=['DELETE'])
def delete_room(room_id):
    room = Room.query.get_or_404(room_id)
    
    # Jihozlar mavjudligini tekshirish
    equipment_count = Equipment.query.filter_by(room_id=room_id).count()
    
    if equipment_count > 0:
        return jsonify({
            'success': False, 
            'has_equipment': True,
            'equipment_count': equipment_count,
            'message': f'Xonada {equipment_count} ta jihoz mavjud. Avval jihozlarni boshqa joyga ko\'chiring yoki o\'chiring.'
        })
    
    # Avval transfer_history jadvalidagi ma'lumotlarni o'chirish
    TransferHistory.query.filter(
        (TransferHistory.from_room_id == room_id) | 
        (TransferHistory.to_room_id == room_id)
    ).delete()
    
    # Keyin xonani o'chirish
    db.session.delete(room)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Xona o\'chirildi'})

# Jihoz transfer qilish
@app.route('/transfer_equipment/<int:equipment_id>', methods=['POST'])
def transfer_equipment(equipment_id):
    equipment = Equipment.query.get_or_404(equipment_id)
    data = request.get_json()
    
    new_room_id = data.get('room_id')
    notes = data.get('notes', '')
    
    if not new_room_id:
        return jsonify({'success': False, 'message': 'Yangi xona tanlanmagan'})
    
    # Yangi xonani tekshirish
    new_room = Room.query.get(new_room_id)
    if not new_room:
        return jsonify({'success': False, 'message': 'Xona topilmadi'})
    
    # Eski xona va yangi xona bir xil bo'lmasligi kerak
    if equipment.room_id == new_room_id:
        return jsonify({'success': False, 'message': 'Jihoz allaqachon bu xonada'})
    
    # Eski ma'lumotlarni saqlash
    old_room_id = equipment.room_id
    old_room_name = equipment.room.name
    old_inv_code = equipment.inv_code
    
    # Jihozni yangi xonaga ko'chirish
    equipment.room_id = new_room_id
    
    # Inventarizatsiya kodini yangilash
    org = equipment.room.organization
    equipment_count = Equipment.query.filter_by(room_id=new_room_id).count()
    new_inv_code = f"{org.name[:3].upper()}-{new_room.name[:3].upper()}-{equipment_count:04d}"
    equipment.inv_code = new_inv_code
    
    # Transfer tarixini saqlash
    transfer_record = TransferHistory(
        equipment_id=equipment_id,
        from_room_id=old_room_id,
        to_room_id=new_room_id,
        old_inv_code=old_inv_code,
        new_inv_code=new_inv_code,
        notes=notes
    )
    
    db.session.add(transfer_record)
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': f'Jihoz "{old_room_name}" dan "{new_room.name}" ga ko\'chirildi',
        'new_inv_code': new_inv_code
    })

# Xonalar ro'yxatini olish (transfer uchun)
@app.route('/get_rooms/<int:organization_id>')
def get_rooms(organization_id):
    rooms = Room.query.filter_by(organization_id=organization_id).all()
    room_list = []
    
    for room in rooms:
        room_data = {
            'id': room.id,
            'name': room.name,
            'floor_name': room.floor.name if room.floor else None
        }
        room_list.append(room_data)
    
    return jsonify(room_list)

# Qavat bo'yicha xonalarni olish
@app.route('/get_rooms_by_floor/<int:floor_id>')
def get_rooms_by_floor(floor_id):
    rooms = Room.query.filter_by(floor_id=floor_id).all()
    room_list = []
    
    for room in rooms:
        room_data = {
            'id': room.id,
            'name': room.name
        }
        room_list.append(room_data)
    
    return jsonify(room_list)

# Jihoz transfer tarixini olish
@app.route('/transfer_history/<int:equipment_id>')
def get_transfer_history(equipment_id):
    transfers = TransferHistory.query.filter_by(equipment_id=equipment_id).order_by(TransferHistory.transfer_date.desc()).all()
    
    history_list = []
    for transfer in transfers:
        history_data = {
            'id': transfer.id,
            'from_room': transfer.from_room.name,
            'to_room': transfer.to_room.name,
            'old_inv_code': transfer.old_inv_code,
            'new_inv_code': transfer.new_inv_code,
            'transfer_date': transfer.transfer_date.strftime('%d.%m.%Y %H:%M'),
            'notes': transfer.notes or ''
        }
        history_list.append(history_data)
    
    return jsonify(history_list)

# Excel fayl yuklash sahifasi
@app.route('/upload_excel/<int:organization_id>')
def upload_excel_page(organization_id):
    organization = Organization.query.get_or_404(organization_id)
    return render_template('upload_excel.html', organization=organization)

# Excel fayl yuklash va ma'lumotlarni import qilish
@app.route('/import_excel/<int:organization_id>', methods=['POST'])
def import_excel(organization_id):
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Fayl tanlanmagan'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Fayl tanlanmagan'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Faqat Excel fayllar (.xlsx, .xls) qabul qilinadi'})
        
        # Excel faylni o'qish
        df = pd.read_excel(file)
        
        # Ustunlarni tekshirish
        required_columns = [
            'P/N(Cihaz Adı)', 'COMPANY(marka)', 'MODEL(Model)', 
            'S/N(Seri Numarası)', 'rangi(Renk)', 'ин/в(Envanter Numarası)',
            'O\'lchov birligi(Miktar)', 'Holati(durum)', 'Kim foydalanayapti(Kim kullanmış)'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False, 
                'message': f'Quyidagi ustunlar topilmadi: {", ".join(missing_columns)}'
            })
        
        # Xona va qavat ma'lumotlarini olish
        room_id = request.form.get('room_id')
        floor_id = request.form.get('floor_id')
        
        if not room_id:
            return jsonify({'success': False, 'message': 'Xona tanlanmagan'})
        
        room = Room.query.get(room_id)
        if not room:
            return jsonify({'success': False, 'message': 'Xona topilmadi'})
        
        # Ma'lumotlarni import qilish
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Ma'lumotlarni tozalash
                name = str(row['P/N(Cihaz Adı)']).strip() if pd.notna(row['P/N(Cihaz Adı)']) else ''
                brand = str(row['COMPANY(marka)']).strip() if pd.notna(row['COMPANY(marka)']) else ''
                model = str(row['MODEL(Model)']).strip() if pd.notna(row['MODEL(Model)']) else ''
                serial_number = str(row['S/N(Seri Numarası)']).strip() if pd.notna(row['S/N(Seri Numarası)']) else ''
                color = str(row['rangi(Renk)']).strip() if pd.notna(row['rangi(Renk)']) else ''
                inv_code = str(row['ин/в(Envanter Numarası)']).strip() if pd.notna(row['ин/в(Envanter Numarası)']) else ''
                quantity = str(row['O\'lchov birligi(Miktar)']).strip() if pd.notna(row['O\'lchov birligi(Miktar)']) else ''
                status = str(row['Holati(durum)']).strip() if pd.notna(row['Holati(durum)']) else 'Active'
                user = str(row['Kim foydalanayapti(Kim kullanmış)']).strip() if pd.notna(row['Kim foydalanayapti(Kim kullanmış)']) else ''
                
                if not name:
                    errors.append(f'Satr {index + 2}: Jihoz nomi bo\'sh')
                    continue
                
                # Inventarizatsiya kodi yaratish
                if not inv_code:
                    equipment_count = Equipment.query.filter_by(room_id=room_id).count() + 1
                    inv_code = f"{room.organization.name[:3].upper()}-{room.name[:3].upper()}-{equipment_count:04d}"
                
                # Jihoz yaratish
                equipment = Equipment(
                    inv_code=inv_code,
                    name=name,
                    category='Import',  # Import qilingan jihozlar uchun
                    brand=brand,
                    model=model,
                    serial_number=serial_number,
                    color=color,
                    status=status,
                    description=f"Miqdor: {quantity}\nFoydalanuvchi: {user}",
                    room_id=room_id
                )
                
                db.session.add(equipment)
                imported_count += 1
                
            except Exception as e:
                errors.append(f'Satr {index + 2}: {str(e)}')
                continue
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{imported_count} ta jihoz muvaffaqiyatli import qilindi',
            'imported_count': imported_count,
            'errors': errors[:10]  # Faqat birinchi 10 ta xatolik
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Import xatoligi: {str(e)}'})

# Excel export funksiyasi
def create_excel_export(organization_id):
    """Tashkilot ma'lumotlarini Excel faylga export qilish"""
    organization = Organization.query.get_or_404(organization_id)
    
    # Yangi Excel workbook yaratish
    wb = Workbook()
    
    # Default sheet ni o'chirish
    wb.remove(wb.active)
    
    # Stil parametrlari
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if organization.has_floors:
        # Qavatli tashkilot uchun
        floors = Floor.query.filter_by(organization_id=organization_id).all()
        
        for floor in floors:
            rooms = Room.query.filter_by(floor_id=floor.id).all()
            
            for room in rooms:
                # Sheet nomi: "Qavat - Xona"
                sheet_name = f"{floor.name} - {room.name}"
                # Excel sheet nomi cheklovlari (31 ta belgi)
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Ma'lumotlarni yozish
                write_room_data_to_sheet(ws, room, header_font, header_fill, border)
    else:
        # Oddiy tashkilot uchun
        rooms = Room.query.filter_by(organization_id=organization_id, floor_id=None).all()
        
        for room in rooms:
            # Sheet nomi: faqat xona nomi
            sheet_name = room.name
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Ma'lumotlarni yozish
            write_room_data_to_sheet(ws, room, header_font, header_fill, border)
    
    # Excel faylni memory ga yozish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def write_room_data_to_sheet(ws, room, header_font, header_fill, border):
    """Xona ma'lumotlarini sheet ga yozish"""
    
    # Sarlavhalar
    headers = [
        'Inv Kode', 'Nomi', 'Kategoriya', 'Brend', 'Model',
        'Seriya Raqami', 'Rang', 'Holat', 'Tavsif'
    ]
    
    # Sarlavhalarni yozish
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Jihozlar ma'lumotlarini yozish
    equipment_list = Equipment.query.filter_by(room_id=room.id).all()
    
    for row, equipment in enumerate(equipment_list, 2):
        data = [
            equipment.inv_code,
            equipment.name,
            equipment.category,
            equipment.brand or '',
            equipment.model or '',
            equipment.serial_number or '',
            equipment.color or '',
            equipment.status,
            equipment.description or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Ustunlarni kengaytirish
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

# Excel export endpoint
@app.route('/export_excel/<int:organization_id>')
def export_excel(organization_id):
    """Tashkilot ma'lumotlarini Excel faylga export qilish"""
    try:
        organization = Organization.query.get_or_404(organization_id)
        
        # Excel fayl yaratish
        excel_file = create_excel_export(organization_id)
        
        # Fayl nomi
        filename = f"{organization.name}_inventarizatsiya.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Xona uchun Excel export funksiyasi
def create_room_excel_export(room_id):
    """Xona ma'lumotlarini Excel faylga export qilish"""
    room = Room.query.get_or_404(room_id)
    
    # Yangi Excel workbook yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = room.name[:31]  # Excel sheet nomi cheklovlari
    
    # Stil parametrlari
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Xona ma'lumotlari
    ws.cell(row=1, column=1, value=f"Xona: {room.name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    if room.floor:
        ws.cell(row=2, column=1, value=f"Qavat: {room.floor.name}")
        ws.cell(row=2, column=1).font = Font(bold=True)
    
    ws.cell(row=3, column=1, value=f"Tashkilot: {room.organization.name}")
    ws.cell(row=3, column=1).font = Font(bold=True)
    
    # Bo'sh qator
    ws.cell(row=4, column=1, value="")
    
    # Sarlavhalar
    headers = [
        'Inv Kode', 'Nomi', 'Kategoriya', 'Brend', 'Model',
        'Seriya Raqami', 'Rang', 'Holat', 'Tavsif'
    ]
    
    # Sarlavhalarni yozish
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Jihozlar ma'lumotlarini yozish
    equipment_list = Equipment.query.filter_by(room_id=room_id).all()
    
    for row, equipment in enumerate(equipment_list, 6):
        data = [
            equipment.inv_code,
            equipment.name,
            equipment.category,
            equipment.brand or '',
            equipment.model or '',
            equipment.serial_number or '',
            equipment.color or '',
            equipment.status,
            equipment.description or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Ustunlarni kengaytirish
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Excel faylni memory ga yozish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# Xona uchun Excel export endpoint
@app.route('/export_room_excel/<int:room_id>')
def export_room_excel(room_id):
    """Xona ma'lumotlarini Excel faylga export qilish"""
    try:
        room = Room.query.get_or_404(room_id)
        
        # Excel fayl yaratish
        excel_file = create_room_excel_export(room_id)
        
        # Fayl nomi
        if room.floor:
            filename = f"{room.organization.name}_{room.floor.name}_{room.name}_inventarizatsiya.xlsx"
        else:
            filename = f"{room.organization.name}_{room.name}_inventarizatsiya.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)
